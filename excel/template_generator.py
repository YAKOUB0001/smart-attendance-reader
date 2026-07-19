# excel_template.py
# -----------------------------------------------------------------------------
# هذه الوحدة مسؤولة عن توليد قالب ملف الإكسل لكشف الحضور والانصراف والراتب.
# تستقبل الشهر والسنة، وتحسب تلقائيًا:
#   - عدد أيام الشهر (28 حتى 31 حسب الشهر والسنة)
#   - اسم اليوم بالعربي لكل تاريخ (الأحد، الإثنين، ...)
# ثم تكتب كل ذلك في ملف Excel بنفس ترتيب وشكل الجدول الأصلي (اتجاه من اليمين لليسار).
#
# تحديث: أُضيف قسم "ملخص الراتب الشهري" أسفل جدول الأيام، يحسب تلقائيًا:
#   - إجمالي ساعات العمل (من جدول الحضور)
#   - الأجر الأساسي = الساعات × سعر الساعة الأساسي
#   - علاوة الأجر   = الساعات × علاوة الساعة
#   - التسبيقات (خصم يُدخل يدويًا)
#   - الأجر النهائي الصافي = (الأجر الأساسي + العلاوة) - التسبيقات
#
# كل الحسابات تُكتب كمعادلات Excel حقيقية (وليست أرقامًا جاهزة)، بحيث لو عدّل
# المحاسب أي خلية (ساعات، سعر، تسبيقات) يُعاد الحساب تلقائيًا.
#
# تحديث آخر: الجدول لم يعد ملاصقًا للحافة اليمنى للورقة. أضفنا عمودين فارغين
# قبل الجدول وعمودين فارغين بعده بنفس العرض تمامًا، بحيث يظهر الجدول في
# منتصف الورقة (وليس عند الحافة) عند فتح الملف في Excel.
# -----------------------------------------------------------------------------

import calendar
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries


# خريطة تحويل رقم اليوم (weekday) إلى اسمه بالعربي
# ملاحظة: في بايثون، date.weekday() تُرجع: الإثنين = 0 ... الأحد = 6
ARABIC_WEEKDAYS = {
    0: "الإثنين",
    1: "الثلاثاء",
    2: "الأربعاء",
    3: "الخميس",
    4: "الجمعة",
    5: "السبت",
    6: "الأحد",
}

ARABIC_MONTHS = {
    1: "يناير", 2: "فبراير", 3: "مارس", 4: "أبريل",
    5: "مايو", 6: "يونيو", 7: "يوليو", 8: "أغسطس",
    9: "سبتمبر", 10: "أكتوبر", 11: "نوفمبر", 12: "ديسمبر",
}

# ألوان وأنماط مستخدمة في القالب
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")   # خلايا تُعبّأ يدويًا
NEUTRAL_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # خلايا محسوبة تلقائيًا
FINAL_FILL = PatternFill(start_color="D9F2E6", end_color="D9F2E6", fill_type="solid")    # خلية الأجر النهائي
TITLE_FILL = PatternFill(start_color="263252", end_color="263252", fill_type="solid")    # عنوان قسم الراتب

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT_MID = Alignment(horizontal="right", vertical="center", wrap_text=True)

# ------------------------------------------------------------------------
# إعدادات توسيط الجدول أفقيًا داخل الورقة
# MARGIN_COLUMNS: عدد الأعمدة الفارغة التي تُضاف قبل الجدول وبعده (بنفس العدد
#                  والعرض على الجانبين) لجعل الجدول يبدو في المنتصف.
# MARGIN_WIDTH:   عرض كل عمود هامش فارغ.
# ------------------------------------------------------------------------
MARGIN_COLUMNS = 2
MARGIN_WIDTH = 10
TABLE_COLUMNS = 5  # عدد أعمدة الجدول الفعلية (التاريخ، الدخول، الخروج، الساعات، الملاحظات)


def _style_merged_range(ws, cell_range, fill=None, font=None, border=None, alignment=None):
    """
    يطبّق التنسيق (تعبئة/خط/حدود/محاذاة) على كل خلايا نطاق مدمج، وليس فقط
    الخلية العلوية اليمنى. مهم لأن openpyxl يخزن القيمة في خلية واحدة فقط،
    لكن الحدود والتعبئة يجب أن تظهر على كامل المنطقة المدمجة بصريًا.
    """
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font
            if border is not None:
                cell.border = border
            if alignment is not None:
                cell.alignment = alignment


def generate_attendance_template(
    month: int,
    year: int,
    employee_name: str = "",
    phone: str = "",
    base_rate: float = 150.0,
    allowance_rate: float = 50.0,
    advances: float = 0.0,
    output_path: str = "attendance_template.xlsx",
):
    """
    يولّد ملف إكسل لكشف حضور وراتب شهر كامل.

    المعاملات:
        month: رقم الشهر (1-12)
        year: السنة (مثال: 2026)
        employee_name: اسم الموظف (اختياري)
        phone: رقم الهاتف (اختياري)
        base_rate: سعر الساعة الأساسي بالدينار (افتراضي 150)
        allowance_rate: علاوة الساعة بالدينار (افتراضي 50)
        advances: قيمة التسبيقات/الخصومات بالدينار (افتراضي 0)
        output_path: مسار حفظ الملف الناتج
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "كشف الحضور والراتب"

    # اتجاه الورقة من اليمين لليسار (RTL) ليطابق اللغة العربية
    ws.sheet_view.rightToLeft = True

    # توسيط الجدول عند الطباعة (أفقيًا) — يجعل المحتوى في منتصف الصفحة المطبوعة
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # ------------------------------------------------------------------
    # إزاحة كل أعمدة الجدول بمقدار MARGIN_COLUMNS حتى يظهر عمود فارغ بنفس
    # العرض قبل الجدول (يمينه) وبعده (يساره)، فيبدو الجدول في المنتصف بدل
    # أن يكون ملاصقًا للحافة اليمنى للورقة.
    # ------------------------------------------------------------------
    def col(idx: int) -> int:
        """رقم العمود الفعلي في الورقة بعد إضافة هامش البداية."""
        return idx + MARGIN_COLUMNS

    def col_letter(idx: int) -> str:
        """حرف العمود الفعلي (مثال: col_letter(1) قد يعطي 'C' بدل 'A')."""
        return get_column_letter(col(idx))

    last_table_col = col(TABLE_COLUMNS)

    # ------------------------------------------------------------------
    # الصف الأول: عنوان الكشف
    # ------------------------------------------------------------------
    ws.merge_cells(start_row=1, start_column=col(1), end_row=1, end_column=last_table_col)
    title_cell = ws.cell(row=1, column=col(1), value=f"كشف الحضور والراتب - شهر {ARABIC_MONTHS[month]} {year}")
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = CENTER

    # ------------------------------------------------------------------
    # الصف الثاني: بيانات الموظف (اسم / هاتف)
    # ------------------------------------------------------------------
    name_label = ws.cell(row=2, column=col(1), value="الاسم:")
    name_label.font = Font(bold=True)
    ws.merge_cells(start_row=2, start_column=col(2), end_row=2, end_column=col(3))
    name_value = ws.cell(row=2, column=col(2), value=employee_name)
    name_value.fill = INPUT_FILL

    phone_label = ws.cell(row=2, column=col(4), value="رقم الهاتف:")
    phone_label.font = Font(bold=True)
    phone_value = ws.cell(row=2, column=col(5), value=phone)
    phone_value.fill = INPUT_FILL

    # ------------------------------------------------------------------
    # صف عنوان جدول الأيام (الصف 4)
    # العمود 1: التاريخ | 2: وقت الدخول | 3: وقت الخروج | 4: مجموع الساعات | 5: الملاحظات
    # ------------------------------------------------------------------
    header_row = 4
    headers = ["التاريخ", "وقت الدخول", "وقت الخروج", "مجموع الساعات", "الملاحظات"]
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col(col_idx), value=title)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # ------------------------------------------------------------------
    # صفوف الأيام: تُحسب تلقائيًا حسب عدد أيام الشهر المُدخل
    # ------------------------------------------------------------------
    days_in_month = calendar.monthrange(year, month)[1]
    first_data_row = header_row + 1

    in_letter = col_letter(2)
    out_letter = col_letter(3)

    for day_num in range(1, days_in_month + 1):
        current_date = date(year, month, day_num)
        day_name = ARABIC_WEEKDAYS[current_date.weekday()]
        row = first_data_row + day_num - 1

        date_cell = ws.cell(row=row, column=col(1), value=f"{day_name}  {month:02d}/{day_num:02d}")
        date_cell.alignment = CENTER
        date_cell.border = THIN_BORDER

        for c_idx in (2, 3):
            c = ws.cell(row=row, column=col(c_idx))
            c.fill = INPUT_FILL
            c.alignment = CENTER
            c.border = THIN_BORDER
            c.number_format = "HH:MM"

        # مجموع الساعات: معادلة تحسب الفرق بين الخروج والدخول (مع دعم عبور منتصف الليل)
        hours_cell = ws.cell(row=row, column=col(4))
        hours_cell.value = (
            f'=IF(AND({in_letter}{row}<>"",{out_letter}{row}<>""),'
            f'MOD({out_letter}{row}-{in_letter}{row},1)*24,"")'
        )
        hours_cell.number_format = "0.00"
        hours_cell.alignment = CENTER
        hours_cell.border = THIN_BORDER

        notes_cell = ws.cell(row=row, column=col(5))
        notes_cell.alignment = CENTER
        notes_cell.border = THIN_BORDER

    # ------------------------------------------------------------------
    # صف إجمالي ساعات العمل الشهرية (نهاية جدول الأيام)
    # ------------------------------------------------------------------
    total_row = first_data_row + days_in_month
    hours_letter = col_letter(4)

    ws.merge_cells(start_row=total_row, start_column=col(1), end_row=total_row, end_column=col(3))
    total_label = ws.cell(row=total_row, column=col(1), value="إجمالي ساعات العمل الشهرية")
    total_label.font = Font(bold=True)
    total_label.alignment = CENTER
    total_label.fill = HEADER_FILL
    total_label.border = THIN_BORDER

    total_value = ws.cell(row=total_row, column=col(4))
    total_value.value = f"=SUM({hours_letter}{first_data_row}:{hours_letter}{total_row - 1})"
    total_value.number_format = "0.00"
    total_value.font = Font(bold=True)
    total_value.alignment = CENTER
    total_value.fill = HEADER_FILL
    total_value.border = THIN_BORDER

    ws.cell(row=total_row, column=col(5)).border = THIN_BORDER
    ws.cell(row=total_row, column=col(5)).fill = HEADER_FILL
    _style_merged_range(ws, f"{col_letter(1)}{total_row}:{col_letter(3)}{total_row}", border=THIN_BORDER)

    # ------------------------------------------------------------------
    # قسم "ملخص الراتب الشهري" — يبدأ بعد سطر فاصل من جدول الأيام
    # ------------------------------------------------------------------
    row = total_row + 2
    value_col_letter = col_letter(4)  # عمود القيم/المعادلات في قسم الملخص (يقابل D سابقًا)

    # عنوان القسم
    ws.merge_cells(start_row=row, start_column=col(1), end_row=row, end_column=col(5))
    title_font = Font(size=13, bold=True, color="FFFFFF")
    ws.cell(row=row, column=col(1), value="ملخص الراتب الشهري")
    _style_merged_range(ws, f"{col_letter(1)}{row}:{col_letter(5)}{row}", fill=TITLE_FILL, font=title_font,
                         border=THIN_BORDER, alignment=CENTER)
    row += 1

    def add_summary_row(label: str, value=None, formula: str = None,
                         editable: bool = False, emphasize: bool = False) -> int:
        """
        يضيف صفًا في قسم ملخص الراتب: أول 3 أعمدة الجدول للتسمية، والعمودان
        الأخيران للقيمة/المعادلة. يُرجع رقم الصف المستخدم حتى يمكن الإشارة
        إليه في معادلات لاحقة.
        """
        nonlocal row
        current_row = row

        # التسمية (يمين، أول 3 أعمدة مدمجة)
        ws.merge_cells(start_row=current_row, start_column=col(1), end_row=current_row, end_column=col(3))
        label_font = Font(bold=True, size=13) if emphasize else Font(bold=False, size=11)
        ws.cell(row=current_row, column=col(1), value=label)
        _style_merged_range(ws, f"{col_letter(1)}{current_row}:{col_letter(3)}{current_row}",
                             border=THIN_BORDER, font=label_font, alignment=RIGHT_MID)

        # القيمة (آخر عمودين مدمجان)
        ws.merge_cells(start_row=current_row, start_column=col(4), end_row=current_row, end_column=col(5))
        value_cell = ws.cell(row=current_row, column=col(4))
        value_cell.value = formula if formula is not None else value
        value_cell.number_format = "#,##0.00"

        if editable:
            fill = INPUT_FILL
        elif emphasize:
            fill = FINAL_FILL
        else:
            fill = NEUTRAL_FILL

        value_font = Font(bold=True, size=13, color="1B5E42") if emphasize else Font(bold=False, size=11)
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thick" if emphasize else "thin"),
        )
        _style_merged_range(ws, f"{col_letter(4)}{current_row}:{col_letter(5)}{current_row}",
                             fill=fill, font=value_font, border=border, alignment=CENTER)

        row += 1
        return current_row

    base_rate_row = add_summary_row("سعر الساعة الأساسي (دج/ساعة)", value=base_rate, editable=True)
    allowance_rate_row = add_summary_row("علاوة الساعة (دج/ساعة)", value=allowance_rate, editable=True)
    total_rate_row = add_summary_row(
        "إجمالي سعر الساعة (دج/ساعة)",
        formula=f"={value_col_letter}{base_rate_row}+{value_col_letter}{allowance_rate_row}",
    )
    hours_row = add_summary_row(
        "إجمالي ساعات العمل (ساعة)",
        formula=f"={value_col_letter}{total_row}",
    )
    base_pay_row = add_summary_row(
        "الأجر الأساسي (دج)",
        formula=f"={value_col_letter}{hours_row}*{value_col_letter}{base_rate_row}",
    )
    allowance_pay_row = add_summary_row(
        "علاوة الأجر (دج)",
        formula=f"={value_col_letter}{hours_row}*{value_col_letter}{allowance_rate_row}",
    )
    advances_row = add_summary_row(
        "التسبيقات / الخصومات (دج)", value=advances, editable=True,
    )
    add_summary_row(
        "الأجر النهائي الصافي (دج)",
        formula=f"=({value_col_letter}{base_pay_row}+{value_col_letter}{allowance_pay_row})-{value_col_letter}{advances_row}",
        emphasize=True,
    )

    # ------------------------------------------------------------------
    # ضبط عرض الأعمدة: أعمدة الجدول بعرضها المعتاد، مع أعمدة هامش فارغة
    # بنفس العرض تمامًا على الجانبين (قبل الجدول وبعده) لتوسيطه بصريًا.
    # ------------------------------------------------------------------
    table_widths = {1: 20, 2: 14, 3: 14, 4: 16, 5: 22}
    for col_idx, width in table_widths.items():
        ws.column_dimensions[get_column_letter(col(col_idx))].width = width

    # أعمدة الهامش قبل الجدول (من العمود 1 إلى MARGIN_COLUMNS)
    for margin_idx in range(1, MARGIN_COLUMNS + 1):
        ws.column_dimensions[get_column_letter(margin_idx)].width = MARGIN_WIDTH

    # أعمدة الهامش بعد الجدول (بنفس العدد والعرض، لضمان تناظر تام)
    after_start = last_table_col + 1
    for offset in range(MARGIN_COLUMNS):
        ws.column_dimensions[get_column_letter(after_start + offset)].width = MARGIN_WIDTH

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    # مثال تجريبي: توليد كشف لشهر أبريل 2026 بسعر ساعة أساسي 150 وعلاوة 50 وتسبيقات 2000
    path = generate_attendance_template(
        month=4, year=2026,
        employee_name="", phone="",
        base_rate=150, allowance_rate=50, advances=2000,
        output_path="output/attendance_2026_04.xlsx",
    )
    print(f"تم إنشاء الملف: {path}")